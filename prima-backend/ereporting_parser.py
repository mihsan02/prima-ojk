"""Parser untuk file e-reporting XLSX (POJK 27/2024).

Column layout diverifikasi cell-by-cell terhadap template resmi:
`Struktur_Data_Laporan_Harian_dan_Bulanan_PAKD.xlsx` dan
`..._Kustodian.xlsx`. Parser tetap defensif (baris/kolom yang kosong atau
tidak sesuai format dilewati dan dicatat di `errors`, bukan menyebabkan
exception) karena upload sebenarnya masih bisa berbeda dari template.
"""

import hashlib
import logging

import openpyxl

logger = logging.getLogger(__name__)

# --- Sheet "LSTAKDKP": Rekapitulasi Aset Keuangan Digital Konsumen dan
# Pedagang Bulanan. (Bukan sheet "2" - itu tidak ada di template resmi;
# workbook PAKD berisi banyak sheet lain seperti LBNP, LBLD, LRA, dst.)
# Header group ada di row 14, subheader di row 15, data mulai row 16.
# Kolom E-AA didahului 4 grup 5-kolom berulang (Posisi Awal Bulan, Nilai
# Penambahan, Nilai Pengurangan, Posisi Akhir Bulan) sebelum kolom target.
PAKD_SHEET2_NAME = "LSTAKDKP"
PAKD_SHEET2_HEADER_ROW = 14
PAKD_SHEET2_DATA_START_ROW = 16
COL_KODE_AKD = 5        # E
COL_NAMA_AKD = 6        # F
COL_JUMLAH_AKD_UNIT = 22     # V  (Posisi Akhir Bulan: Jumlah Unit)
COL_KONSUMEN_UNIT = 23       # W  (Posisi Akhir Bulan: milik Konsumen)
COL_PEDAGANG_UNIT = 24       # X  (Posisi Akhir Bulan: milik Pedagang)
COL_KONSUMEN_DI_PEDAGANG = 25  # Y  (Posisi Akhir Bulan: Konsumen di Pedagang)
COL_KONSUMEN_DI_PTP = 26       # Z  (Posisi Akhir Bulan: Konsumen di PTP)
COL_HARGA_PER_UNIT = 27        # AA (Harga Penutupan Per Unit)

LBNP_SHEET_NAME = "LBNP"
LRA_SHEET_NAME = "LRA"
LABEL_COL = 3   # C
VALUE_COL = 5   # E
LABEL_VALUE_START_ROW = 11

KUSTODIAN_SHEET_NAME = "LPWAKD"
KUSTODIAN_DATA_START_ROW = 13
KUSTODIAN_HEADER_SCAN_ROWS = 14  # scan rows 1..14 looking for header labels (real header is row 11)
# Fallback columns, verified against the real template (row 11 header,
# cols E-J). Used only if dynamic header detection can't confirm a match.
KUSTODIAN_FALLBACK_COLS = {
    "alamat": 5,          # E
    "provider": 6,        # F
    "network": 7,         # G
    "nama_pedagang": 8,   # H
    "nilai_akd_idr": 9,   # I
    "keterangan": 10,     # J
}
KUSTODIAN_HEADER_KEYWORDS = {
    "alamat": ["alamat wallet", "alamat"],
    "provider": ["nama provider", "provider"],
    "network": ["network", "jaringan"],
    "nama_pedagang": ["nama pedagang", "pedagang"],
    "nilai_akd_idr": ["nilai akd", "nilai"],
    "keterangan": ["keterangan"],
}


def _safe_float(val):
    """Safely cast any value to float. Returns 0.0 on failure."""
    if val is None:
        return 0.0
    if isinstance(val, bool):
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip()
    if not s:
        return 0.0
    s = s.replace("Rp", "").replace("rp", "").replace(",", "").strip()
    try:
        return float(s)
    except (ValueError, TypeError):
        return 0.0


def _compute_hash(filepath):
    """SHA-256 hash of file contents."""
    h = hashlib.sha256()
    with open(filepath, "rb") as f:
        for chunk in iter(lambda: f.read(65536), b""):
            h.update(chunk)
    return h.hexdigest()


def _cell(ws, row, col):
    try:
        return ws.cell(row=row, column=col).value
    except Exception:
        return None


def _parse_sheet2(ws, errors):
    breakdown = []
    row = PAKD_SHEET2_DATA_START_ROW
    empty_streak = 0
    while empty_streak < 5:
        kode_akd = _cell(ws, row, COL_KODE_AKD)
        if kode_akd is None or str(kode_akd).strip() == "":
            empty_streak += 1
            row += 1
            continue
        empty_streak = 0
        breakdown.append({
            "kode_akd": str(kode_akd).strip(),
            "nama_akd": str(_cell(ws, row, COL_NAMA_AKD) or "").strip(),
            "posisi_akhir_unit": _safe_float(_cell(ws, row, COL_JUMLAH_AKD_UNIT)),
            "konsumen_unit": _safe_float(_cell(ws, row, COL_KONSUMEN_UNIT)),
            "pedagang_unit": _safe_float(_cell(ws, row, COL_PEDAGANG_UNIT)),
            "konsumen_di_pedagang_unit": _safe_float(_cell(ws, row, COL_KONSUMEN_DI_PEDAGANG)),
            "konsumen_di_ptp_unit": _safe_float(_cell(ws, row, COL_KONSUMEN_DI_PTP)),
            "harga_per_unit_idr": _safe_float(_cell(ws, row, COL_HARGA_PER_UNIT)),
        })
        row += 1
    if not breakdown:
        errors.append(f"Sheet '{PAKD_SHEET2_NAME}': tidak ada baris data aset ditemukan")
    return breakdown


def _scan_label_value(ws, label_map):
    """Scan the label column for substrings (case-insensitive), read value from the value column.

    label_map: dict of {result_key: (must_contain[], must_not_contain[])}
    Returns dict of {result_key: float value}, defaulting to 0.0 when not found.
    """
    result = {key: 0.0 for key in label_map}
    max_row = ws.max_row or 0
    for row in range(LABEL_VALUE_START_ROW, max_row + 1):
        label = _cell(ws, row, LABEL_COL)
        if label is None:
            continue
        label_lower = str(label).strip().lower()
        if not label_lower:
            continue
        for key, (must_contain, must_not_contain) in label_map.items():
            if any(term not in label_lower for term in must_contain):
                continue
            if any(term in label_lower for term in must_not_contain):
                continue
            result[key] = _safe_float(_cell(ws, row, VALUE_COL))
    return result


def _parse_lbnp(ws):
    label_map = {
        "persediaan_akd_idr": (["persediaan aset keuangan digital"], []),
        "simpanan_pedagang_idr": (["simpanan milik pedagang"], []),
        "ekuitas_idr": (["jumlah ekuitas"], []),
        "liabilitas_konsumen_idr": (["liabilitas kepada konsumen"], ["penggantian"]),
    }
    return _scan_label_value(ws, label_map)


def _parse_lra(ws):
    label_map = {
        "titipan_akd_di_pedagang_idr": (["pada pedagang"], ["titipan"]),
        "titipan_akd_di_ptp_idr": (["pada pengelola tempat penyimp"], []),
    }
    return _scan_label_value(ws, label_map)


def parse_pakd_ereporting(filepath: str) -> dict:
    """Parse PAKD e-reporting XLSX (Sheet LSTAKDKP + LBNP + LRA).

    Returns {file_hash, aset_breakdown: [...], balance_sheet: {...},
             rekening_administratif: {...}, errors: [...]}
    """
    errors = []
    aset_breakdown = []
    balance_sheet = {}
    rekening_administratif = {}
    file_hash = None

    try:
        file_hash = _compute_hash(filepath)
    except Exception as e:
        errors.append(f"Gagal menghitung hash file: {e}")

    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
    except Exception as e:
        errors.append(f"Gagal membuka file XLSX: {e}")
        return {
            "file_hash": file_hash,
            "aset_breakdown": aset_breakdown,
            "balance_sheet": balance_sheet,
            "rekening_administratif": rekening_administratif,
            "errors": errors,
        }

    if PAKD_SHEET2_NAME in wb.sheetnames:
        aset_breakdown = _parse_sheet2(wb[PAKD_SHEET2_NAME], errors)
    else:
        logger.warning("Sheet '%s' tidak ditemukan di %s", PAKD_SHEET2_NAME, filepath)
        errors.append(f"Sheet '{PAKD_SHEET2_NAME}' tidak ditemukan")

    if LBNP_SHEET_NAME in wb.sheetnames:
        balance_sheet = _parse_lbnp(wb[LBNP_SHEET_NAME])
    else:
        logger.warning("Sheet '%s' tidak ditemukan di %s", LBNP_SHEET_NAME, filepath)
        errors.append(f"Sheet '{LBNP_SHEET_NAME}' tidak ditemukan")

    if LRA_SHEET_NAME in wb.sheetnames:
        rekening_administratif = _parse_lra(wb[LRA_SHEET_NAME])
    else:
        logger.warning("Sheet '%s' tidak ditemukan di %s", LRA_SHEET_NAME, filepath)
        errors.append(f"Sheet '{LRA_SHEET_NAME}' tidak ditemukan")

    return {
        "file_hash": file_hash,
        "aset_breakdown": aset_breakdown,
        "balance_sheet": balance_sheet,
        "rekening_administratif": rekening_administratif,
        "errors": errors,
    }


def _detect_kustodian_header_cols(ws):
    """Try to locate LPWAKD columns by scanning header rows for known labels.

    Returns a dict {field: column_index} on success, or None if no header
    row confidently matched (caller should fall back to KUSTODIAN_FALLBACK_COLS).
    """
    max_col = ws.max_column or 12
    for row in range(1, KUSTODIAN_HEADER_SCAN_ROWS + 1):
        found = {}
        for col in range(1, max_col + 1):
            val = _cell(ws, row, col)
            if val is None:
                continue
            val_lower = str(val).strip().lower()
            if not val_lower:
                continue
            for field, keywords in KUSTODIAN_HEADER_KEYWORDS.items():
                if field in found:
                    continue
                if any(kw in val_lower for kw in keywords):
                    found[field] = col
        # Require at least "alamat" + one more field to trust this row as header
        if "alamat" in found and len(found) >= 3:
            return found
    return None


def parse_kustodian_wallet_report(filepath: str) -> dict:
    """Parse Kustodian wallet report XLSX (Sheet LPWAKD).

    Returns {file_hash, wallets: [...], errors: [...]}
    """
    errors = []
    wallets = []
    file_hash = None

    try:
        file_hash = _compute_hash(filepath)
    except Exception as e:
        errors.append(f"Gagal menghitung hash file: {e}")

    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
    except Exception as e:
        errors.append(f"Gagal membuka file XLSX: {e}")
        return {"file_hash": file_hash, "wallets": wallets, "errors": errors}

    if KUSTODIAN_SHEET_NAME not in wb.sheetnames:
        logger.warning("Sheet '%s' tidak ditemukan di %s", KUSTODIAN_SHEET_NAME, filepath)
        errors.append(f"Sheet '{KUSTODIAN_SHEET_NAME}' tidak ditemukan")
        return {"file_hash": file_hash, "wallets": wallets, "errors": errors}

    ws = wb[KUSTODIAN_SHEET_NAME]
    cols = _detect_kustodian_header_cols(ws)
    if cols is None:
        cols = KUSTODIAN_FALLBACK_COLS
        errors.append(
            f"Sheet '{KUSTODIAN_SHEET_NAME}': header kolom tidak terdeteksi, "
            "menggunakan posisi kolom default (E-J)"
        )

    row = KUSTODIAN_DATA_START_ROW
    empty_streak = 0
    while empty_streak < 5:
        alamat = _cell(ws, row, cols["alamat"])
        if alamat is None or str(alamat).strip() == "":
            empty_streak += 1
            row += 1
            continue
        empty_streak = 0
        wallets.append({
            "alamat": str(alamat).strip(),
            "provider": str(_cell(ws, row, cols["provider"]) or "").strip(),
            "network": str(_cell(ws, row, cols["network"]) or "").strip(),
            "nama_pedagang": str(_cell(ws, row, cols["nama_pedagang"]) or "").strip(),
            "nilai_akd_idr": _safe_float(_cell(ws, row, cols["nilai_akd_idr"])),
            "keterangan": str(_cell(ws, row, cols["keterangan"]) or "").strip(),
        })
        row += 1

    if not wallets:
        errors.append(f"Sheet '{KUSTODIAN_SHEET_NAME}': tidak ada baris wallet ditemukan")

    return {"file_hash": file_hash, "wallets": wallets, "errors": errors}
