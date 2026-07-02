"""
Sprint 3: E-reporting parser + upload/confirm endpoint test suite.
Tests parser logic directly (no DB) and endpoint contracts (mocked DB).
"""
import os
import time
import json
import tempfile

import openpyxl
import pytest
from unittest.mock import patch, MagicMock

os.environ.setdefault('ADMIN_TOKEN', 'test-token-prima')
os.environ.setdefault('SUPABASE_JWT_SECRET', 'prima-test-jwt-secret-for-pytest-1234')

from app import app as flask_app
from ereporting_parser import (
    parse_pakd_ereporting,
    parse_kustodian_wallet_report,
    _safe_float,
    _compute_hash,
)

JWT_SECRET = os.environ['SUPABASE_JWT_SECRET']

MOCK_PROFILES = {
    'admin-001': {'role': 'super_admin', 'entity_type': None, 'entity_id': None, 'display_name': 'Super Admin'},
    'pengawas-001': {'role': 'pengawas', 'entity_type': None, 'entity_id': None, 'display_name': 'Pengawas IAKD'},
}


def _mock_fetch_profile(user_id):
    return MOCK_PROFILES.get(user_id)


def _make_token(user_id='admin-001'):
    import jwt as pyjwt
    now = int(time.time())
    payload = {'sub': user_id, 'email': f'{user_id}@test.com', 'aud': 'authenticated', 'iat': now, 'exp': now + 3600}
    return pyjwt.encode(payload, JWT_SECRET, algorithm='HS256')


@pytest.fixture
def client():
    flask_app.config['TESTING'] = True
    flask_app.config['PROPAGATE_EXCEPTIONS'] = True
    with flask_app.test_client() as c:
        yield c


def _make_mock_conn(query_results=None):
    mock_conn = MagicMock()
    mock_cur = MagicMock()
    mock_conn.cursor.return_value = mock_cur
    if query_results:
        mock_cur.fetchone.side_effect = query_results.get('fetchone', [None])
        mock_cur.fetchall.side_effect = query_results.get('fetchall', [[]])
    else:
        mock_cur.fetchone.return_value = None
        mock_cur.fetchall.return_value = []
    return mock_conn


def _create_pakd_xlsx(aset_data=None, balance_sheet=None, lra_data=None,
                       include_sheets=('LSTAKDKP', 'LBNP', 'LRA')):
    """Build a PAKD XLSX fixture at coordinates verified cell-by-cell against
    the real POJK 27/2024 template (LSTAKDKP row 14/15 header, row 16 data,
    cols E/F/V/W/X/Y/Z/AA; LBNP/LRA label col C, value col E, from row 11).

    aset_data: list of (kode, nama, unit, konsumen, pedagang, di_pedagang, di_ptp, harga)
               tuples; any field may be None to leave that cell blank.
    """
    wb = openpyxl.Workbook()
    if 'LSTAKDKP' in include_sheets:
        ws = wb.create_sheet('LSTAKDKP')
        ws['D14'] = 'Nomor Baris'
        ws['E14'] = 'Kode Aset Keuangan Digital'
        ws['F14'] = 'Nama Aset Keuangan Digital'
        ws['V14'] = 'Posisi Akhir Bulan'
        if aset_data:
            for i, (kode, nama, unit, konsumen, pedagang, di_pedagang, di_ptp, harga) in enumerate(aset_data):
                r = 16 + i
                ws[f'D{r}'] = i + 1
                if kode is not None:
                    ws[f'E{r}'] = kode
                if nama is not None:
                    ws[f'F{r}'] = nama
                if unit is not None:
                    ws[f'V{r}'] = unit
                if konsumen is not None:
                    ws[f'W{r}'] = konsumen
                if pedagang is not None:
                    ws[f'X{r}'] = pedagang
                if di_pedagang is not None:
                    ws[f'Y{r}'] = di_pedagang
                if di_ptp is not None:
                    ws[f'Z{r}'] = di_ptp
                if harga is not None:
                    ws[f'AA{r}'] = harga
    if 'LBNP' in include_sheets:
        ws_lbnp = wb.create_sheet('LBNP')
        if balance_sheet:
            for i, (label, value) in enumerate(balance_sheet):
                r = 11 + i
                ws_lbnp[f'C{r}'] = label
                ws_lbnp[f'E{r}'] = value
    if 'LRA' in include_sheets:
        ws_lra = wb.create_sheet('LRA')
        if lra_data:
            for i, (label, value) in enumerate(lra_data):
                r = 11 + i
                ws_lra[f'C{r}'] = label
                ws_lra[f'E{r}'] = value
    if 'Sheet' in wb.sheetnames and len(wb.sheetnames) > 1:
        del wb['Sheet']
    tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
    wb.save(tmp.name)
    tmp.close()
    return tmp.name


def _create_kustodian_xlsx(wallet_rows=None, with_header=True, include_sheet=True):
    """Build a Kustodian LPWAKD XLSX fixture at coordinates verified
    cell-by-cell against the real template (header row 11, data row 13,
    cols E-J).

    wallet_rows: list of (addr, provider, net, pedagang, nilai, ket) tuples.
    """
    wb = openpyxl.Workbook()
    if include_sheet:
        ws = wb.create_sheet('LPWAKD')
        if with_header:
            ws['E11'] = 'Alamat Wallet yang Digunakan'
            ws['F11'] = 'Nama Provider'
            ws['G11'] = 'Network'
            ws['H11'] = 'Nama Pedagang'
            ws['I11'] = 'Nilai Aset Keuangan Digital (Rp)'
            ws['J11'] = 'Keterangan'
        if wallet_rows:
            for i, (addr, provider, net, pedagang, nilai, ket) in enumerate(wallet_rows):
                r = 13 + i
                ws[f'C{r}'] = i + 1
                ws[f'E{r}'] = addr
                ws[f'F{r}'] = provider
                ws[f'G{r}'] = net
                ws[f'H{r}'] = pedagang
                ws[f'I{r}'] = nilai
                ws[f'J{r}'] = ket
    if 'Sheet' in wb.sheetnames and len(wb.sheetnames) > 1:
        del wb['Sheet']
    tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
    wb.save(tmp.name)
    tmp.close()
    return tmp.name


# --- One canonical PAKD fixture reused across several assertions ---
_ASET_DATA = [
    ('AKD001', 'Token Alpha', 1000, 700, 300, 200, 100, 50000),
    ('AKD002', 'Token Beta', 500, 400, 100, 50, 50, 20000),
    (None, None, None, None, None, None, None, None),
]
_LBNP_DATA = [
    ('Persediaan Aset Keuangan Digital', 1_000_000),
    ('Simpanan milik Pedagang', 2_000_000),
    ('JUMLAH EKUITAS', 5_000_000),
    ('Liabilitas kepada Konsumen', 3_000_000),
    ('Liabilitas Penggantian kepada Konsumen', 999_999),
]
_LRA_DATA = [
    ('Titipan AKD Konsumen Pada Pedagang', 111),
    ('AKD Milik Konsumen Pada Pedagang', 222),
    ('Pada Pengelola Tempat Penyimpanan', 333),
]


class TestPakdParser:
    def test_parse_pakd_sheet2(self):
        path = _create_pakd_xlsx(aset_data=_ASET_DATA)
        result = parse_pakd_ereporting(path)
        breakdown = result['aset_breakdown']
        assert len(breakdown) == 2
        assert breakdown[0]['kode_akd'] == 'AKD001'
        assert breakdown[0]['nama_akd'] == 'Token Alpha'
        assert breakdown[0]['konsumen_di_pedagang_unit'] == 200.0
        assert breakdown[0]['konsumen_di_ptp_unit'] == 100.0
        assert breakdown[0]['pedagang_unit'] == 300.0
        assert breakdown[0]['harga_per_unit_idr'] == 50000.0

    def test_parse_pakd_lbnp(self):
        path = _create_pakd_xlsx(balance_sheet=_LBNP_DATA)
        result = parse_pakd_ereporting(path)
        bs = result['balance_sheet']
        assert bs['persediaan_akd_idr'] == 1_000_000.0
        assert bs['simpanan_pedagang_idr'] == 2_000_000.0
        assert bs['ekuitas_idr'] == 5_000_000.0
        # "Liabilitas kepada Konsumen" must exclude the "Penggantian" variant row
        assert bs['liabilitas_konsumen_idr'] == 3_000_000.0

    def test_parse_pakd_lra(self):
        path = _create_pakd_xlsx(lra_data=_LRA_DATA)
        result = parse_pakd_ereporting(path)
        admin = result['rekening_administratif']
        # "Titipan ... Pada Pedagang" row is excluded (contains "titipan");
        # the non-titipan "Pada Pedagang" row is what gets matched
        assert admin['titipan_akd_di_pedagang_idr'] == 222.0
        assert admin['titipan_akd_di_ptp_idr'] == 333.0

    def test_parse_pakd_missing_sheets(self):
        path = _create_pakd_xlsx(include_sheets=())
        result = parse_pakd_ereporting(path)
        assert result['aset_breakdown'] == []
        assert result['balance_sheet'] == {}
        assert result['rekening_administratif'] == {}
        assert len(result['errors']) == 3
        assert any('LSTAKDKP' in e for e in result['errors'])
        assert any('LBNP' in e for e in result['errors'])
        assert any('LRA' in e for e in result['errors'])

    def test_parse_pakd_empty_rows_skipped(self):
        # third row in _ASET_DATA has no kode_akd -> must not appear in output
        path = _create_pakd_xlsx(aset_data=_ASET_DATA)
        result = parse_pakd_ereporting(path)
        assert len(result['aset_breakdown']) == 2
        assert all(row['kode_akd'] for row in result['aset_breakdown'])

    def test_parse_pakd_matches_real_template_layout(self):
        """Regression guard: verified against real POJK 27/2024 PAKD template
        coordinates (LSTAKDKP, header row 14/15, data row 16, cols E-AA).
        Fails loudly if the column offsets are ever silently reintroduced wrong."""
        path = _create_pakd_xlsx(aset_data=[
            ('BTC', 'Bitcoin', 10, 7, 3, 2, 1, 950_000_000),
        ])
        result = parse_pakd_ereporting(path)
        assert result['errors'] == []
        assert len(result['aset_breakdown']) == 1
        assert result['aset_breakdown'][0]['kode_akd'] == 'BTC'
        assert result['aset_breakdown'][0]['harga_per_unit_idr'] == 950_000_000

    def test_parse_pakd_corrupt_file_returns_error(self):
        tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        tmp.write(b'not a real xlsx file')
        tmp.close()
        result = parse_pakd_ereporting(tmp.name)
        assert result['aset_breakdown'] == []
        assert any('Gagal membuka file XLSX' in e for e in result['errors'])


class TestKustodianParser:
    def test_parse_kustodian_lpwakd(self):
        rows = [
            ('0xABC123', 'MetaMask', 'ethereum', 'Pedagang A', 1_000_000, 'aktif'),
            ('bc1qxyz', 'Wallet BTC', 'bitcoin', 'Pedagang B', 2_000_000, ''),
        ]
        path = _create_kustodian_xlsx(wallet_rows=rows)
        result = parse_kustodian_wallet_report(path)
        wallets = result['wallets']
        assert len(wallets) == 2
        assert wallets[0]['alamat'] == '0xABC123'
        assert wallets[0]['provider'] == 'MetaMask'
        assert wallets[0]['network'] == 'ethereum'
        assert wallets[0]['nilai_akd_idr'] == 1_000_000.0

    def test_parse_kustodian_missing_lpwakd(self):
        path = _create_kustodian_xlsx(include_sheet=False)
        result = parse_kustodian_wallet_report(path)
        assert result['wallets'] == []
        assert any('LPWAKD' in e for e in result['errors'])

    def test_parse_kustodian_no_header_falls_back(self):
        rows = [('0xFallback', 'ProviderX', 'ethereum', 'PedagangX', 999, 'note')]
        path = _create_kustodian_xlsx(wallet_rows=rows, with_header=False)
        result = parse_kustodian_wallet_report(path)
        assert len(result['wallets']) == 1
        assert result['wallets'][0]['alamat'] == '0xFallback'
        assert any('header kolom tidak terdeteksi' in e for e in result['errors'])


class TestSafeFloatAndHash:
    def test_safe_float_edge_cases(self):
        assert _safe_float(None) == 0.0
        assert _safe_float('') == 0.0
        assert _safe_float('abc') == 0.0
        assert _safe_float(-123.45) == -123.45
        assert _safe_float('-123.45') == -123.45
        assert _safe_float('Rp 1,000') == 1000.0
        assert _safe_float(True) == 0.0
        assert _safe_float(42) == 42.0

    def test_file_hash_deterministic(self):
        path = _create_pakd_xlsx(aset_data=_ASET_DATA)
        h1 = _compute_hash(path)
        h2 = _compute_hash(path)
        assert h1 == h2
        assert len(h1) == 64  # sha256 hex digest


class TestUploadEndpoint:
    def test_upload_requires_auth(self, client):
        resp = client.post('/api/upload-ereporting', data={'type': 'pakd'})
        assert resp.status_code == 401

    def test_upload_endpoint_requires_super_admin(self, client):
        token = _make_token('pengawas-001')
        with patch('auth._fetch_user_profile', side_effect=_mock_fetch_profile):
            resp = client.post('/api/upload-ereporting',
                                headers={'Authorization': f'Bearer {token}'},
                                data={'type': 'pakd', 'entity_id': 'PAKD-DEMO-001', 'periode': '2026-06'})
            assert resp.status_code == 403

    def test_upload_rejects_non_xlsx(self, client):
        token = _make_token('admin-001')
        import io
        with patch('auth._fetch_user_profile', side_effect=_mock_fetch_profile):
            resp = client.post('/api/upload-ereporting',
                                headers={'Authorization': f'Bearer {token}'},
                                data={'file': (io.BytesIO(b'hello'), 'test.txt'),
                                      'type': 'pakd', 'entity_id': 'PAKD-DEMO-001', 'periode': '2026-06'},
                                content_type='multipart/form-data')
            assert resp.status_code == 400
            assert '.xlsx' in resp.get_json()['message']

    def test_upload_rejects_invalid_type(self, client):
        token = _make_token('admin-001')
        import io
        with patch('auth._fetch_user_profile', side_effect=_mock_fetch_profile):
            resp = client.post('/api/upload-ereporting',
                                headers={'Authorization': f'Bearer {token}'},
                                data={'file': (io.BytesIO(b'x'), 'test.xlsx'),
                                      'type': 'bogus', 'entity_id': 'X', 'periode': '2026-06'},
                                content_type='multipart/form-data')
            assert resp.status_code == 400

    def test_upload_rejects_missing_entity(self, client):
        token = _make_token('admin-001')
        mock_conn = _make_mock_conn({'fetchone': [None]})
        import io
        with patch('auth._fetch_user_profile', side_effect=_mock_fetch_profile), \
             patch('app._get_db_conn', return_value=mock_conn), \
             patch('app._return_db_conn'):
            resp = client.post('/api/upload-ereporting',
                                headers={'Authorization': f'Bearer {token}'},
                                data={'file': (io.BytesIO(b'x'), 'test.xlsx'),
                                      'type': 'pakd', 'entity_id': 'PAKD-NONE', 'periode': '2026-06'},
                                content_type='multipart/form-data')
            assert resp.status_code == 404

    def test_upload_returns_preview_for_valid_entity(self, client):
        token = _make_token('admin-001')
        mock_conn = _make_mock_conn({'fetchone': [('PAKD-DEMO-001',)]})
        path = _create_pakd_xlsx(aset_data=_ASET_DATA)
        with patch('auth._fetch_user_profile', side_effect=_mock_fetch_profile), \
             patch('app._get_db_conn', return_value=mock_conn), \
             patch('app._return_db_conn'):
            with open(path, 'rb') as f:
                resp = client.post('/api/upload-ereporting',
                                    headers={'Authorization': f'Bearer {token}'},
                                    data={'file': (f, 'test.xlsx'),
                                          'type': 'pakd', 'entity_id': 'PAKD-DEMO-001', 'periode': '2026-06'},
                                    content_type='multipart/form-data')
            assert resp.status_code == 200
            body = resp.get_json()
            assert body['status'] == 'preview'
            assert len(body['data']['aset_breakdown']) == 2


class TestConfirmEndpoint:
    def test_confirm_requires_super_admin(self, client):
        token = _make_token('pengawas-001')
        with patch('auth._fetch_user_profile', side_effect=_mock_fetch_profile):
            resp = client.post('/api/confirm-ereporting',
                                headers={'Authorization': f'Bearer {token}'},
                                json={'entity_id': 'PAKD-DEMO-001', 'periode': '2026-06', 'report_type': 'pakd'})
            assert resp.status_code == 403

    def test_confirm_empty_body_returns_400(self, client):
        token = _make_token('admin-001')
        with patch('auth._fetch_user_profile', side_effect=_mock_fetch_profile):
            resp = client.post('/api/confirm-ereporting',
                                headers={'Authorization': f'Bearer {token}'},
                                json={})
            assert resp.status_code == 400

    def test_confirm_saves_to_db(self, client):
        token = _make_token('admin-001')
        mock_conn = _make_mock_conn()
        with patch('auth._fetch_user_profile', side_effect=_mock_fetch_profile), \
             patch('app._get_db_conn', return_value=mock_conn), \
             patch('app._return_db_conn'), \
             patch('app.write_audit') as mock_audit:
            resp = client.post('/api/confirm-ereporting',
                                headers={'Authorization': f'Bearer {token}'},
                                json={
                                    'entity_id': 'PAKD-DEMO-001', 'periode': '2026-06', 'report_type': 'pakd',
                                    'file_hash': 'abc123',
                                    'aset_breakdown': [{'konsumen_di_pedagang_unit': 10, 'konsumen_di_ptp_unit': 5,
                                                         'pedagang_unit': 2, 'harga_per_unit_idr': 1000}],
                                    'balance_sheet': {'ekuitas_idr': 5000000},
                                })
            assert resp.status_code == 200
            body = resp.get_json()
            assert body['status'] == 'confirmed'
            assert body['entity_id'] == 'PAKD-DEMO-001'
            assert mock_audit.called
            # derived values computed correctly: 10*1000, 5*1000, 2*1000
            sql, params = mock_conn.cursor.return_value.execute.call_args[0]
            assert 10000.0 in params and 5000.0 in params and 2000.0 in params

    def test_confirm_upsert_overwrites(self, client):
        """Re-confirming the same entity+periode+report_type must hit the
        ON CONFLICT upsert branch (single INSERT statement handles both)."""
        token = _make_token('admin-001')
        mock_conn = _make_mock_conn()
        with patch('auth._fetch_user_profile', side_effect=_mock_fetch_profile), \
             patch('app._get_db_conn', return_value=mock_conn), \
             patch('app._return_db_conn'), \
             patch('app.write_audit'):
            payload = {
                'entity_id': 'PAKD-DEMO-001', 'periode': '2026-06', 'report_type': 'pakd',
                'aset_breakdown': [], 'balance_sheet': {},
            }
            resp1 = client.post('/api/confirm-ereporting',
                                 headers={'Authorization': f'Bearer {token}'}, json=payload)
            resp2 = client.post('/api/confirm-ereporting',
                                 headers={'Authorization': f'Bearer {token}'}, json=payload)
            assert resp1.status_code == 200
            assert resp2.status_code == 200
            sql = mock_conn.cursor.return_value.execute.call_args[0][0]
            assert 'ON CONFLICT' in sql
            assert 'DO UPDATE SET' in sql

    def test_confirm_kustodian_report(self, client):
        token = _make_token('admin-001')
        mock_conn = _make_mock_conn()
        with patch('auth._fetch_user_profile', side_effect=_mock_fetch_profile), \
             patch('app._get_db_conn', return_value=mock_conn), \
             patch('app._return_db_conn'), \
             patch('app.write_audit'):
            resp = client.post('/api/confirm-ereporting',
                                headers={'Authorization': f'Bearer {token}'},
                                json={'entity_id': 'KUST-001', 'periode': '2026-06', 'report_type': 'kustodian',
                                      'wallets': [{'alamat': '0xabc'}]})
            assert resp.status_code == 200
            assert resp.get_json()['status'] == 'confirmed'
